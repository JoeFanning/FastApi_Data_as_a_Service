import os
import httpx
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

print("--- Starting Premium DaaS Data Pipeline (Header Inspection Mode) ---")

# This function makes an outgoing call to another company's website (like the RentCast API) to ask for data and converts it into
# a standard Python dictionary or list that your script can easily read.
def fetch_raw_api_sample(city_name: str):
    """
    Fetches raw property objects directly from the live RentCast API.
    """
    target_city = city_name.strip() if city_name else "Dallas"
    print(f"Connecting to live database stream to inspect data headers for: {target_city}...")

    url = "https://rentcast.io"

    query_params = {
        "city": target_city,
        "state": "TX",
        "status": "Active",
        "limit": 5
    }

    api_key = os.getenv("RENTCAST_REAL_ESTATE_API_KEY")
    if not api_key:
        print("CRITICAL ERROR: RENTCAST_REAL_ESTATE_API_KEY environment variable is missing!")
        return []

    headers = {
        "Accept": "application/json",
        "X-Api-Key": api_key
    }

    try:
        # This line makes an outgoing call to another company's website (like the RentCast API) to ask for data.
        response = httpx.get(url, headers=headers, params=query_params, timeout=20)
        if response.status_code == 200:
            # This line of code takes the raw data your program just received from the internet and converts it into
            # a standard Python dictionary or list that your script can easily read.
            print(response.json())
            return response.json()
        else:
            print(f"API Warning: Status {response.status_code}. Details: {response.text}")
            return []
    except Exception as e:
        print(f"Network error: {e}")
        return []


def extract_raw_headers(raw_properties):
    # The API Server can return json data as a list [] or as a dictionary {} the code below handles both senarios .
    if isinstance(raw_properties, dict):
        print(f"API returned a Dictionary structure. Available top-level keys: {list(raw_properties.keys())}")
        actual_list = raw_properties.get("listings") or raw_properties.get("data") or []
        raw_properties = actual_list if isinstance(actual_list, list) else [raw_properties]
        print(f" Top level keys::::::::  {actual_list} ")  # Fixed missing parentheses for Python 3 print syntax

    # If the data is not a list or dictionary or is corrupt or nothing is returned this code will handle it
    if not raw_properties or not isinstance(raw_properties, list) or len(raw_properties) == 0:
        print("Data stream empty or malformed. Ensure your key is active in GitHub Secrets.")
        return pd.DataFrame(
            [{"Available API Column Keys": "ERROR", "Sample Raw Data Example": "No data returned from API."}])

    # This takes the entire list of house dictionaries and converts it into a full table instantly
    all_houses_df = pd.DataFrame(raw_properties)

    # Return the dataframe so your web server can read it and display it on the home page
    return all_houses_df


# his tells Python, "Only run the code below if someone directly executes this specific script file."
# It prevents the code from running accidentally if this file is imported into another script.
if __name__ == "__main__":
    # Looks for a system configuration setting named "FILTER_TEXT". If it doesn't find one, it safely
    # defaults to using "Dallas" as the target city.
    input_location = os.getenv("FILTER_TEXT", "Dallas")

    # 1. Harvests json data from server to a dictionary into a variable input_location
    # Then rename it raw_sample
    raw_sample = fetch_raw_api_sample(input_location)

    # 2. Extract schemas
    headers_df = extract_raw_headers(raw_sample)

    file_name = "Premium_Undervalued_Deals.xlsx"

    # 3. Export & style spreadsheet report
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        headers_df.to_excel(writer, sheet_name="API Keys Blueprint", index=False)


        workbook = writer.book
        worksheet = writer.sheets["API Keys Blueprint"]

        # Design system styles
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")  # Purple
        data_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")  # Light grey

        # Single-pass layout formatting loop
        for col_idx, col in enumerate(worksheet.columns, 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)

            for row_idx, cell in enumerate(col, 1):
                # Calculate text padding dynamically
                max_len = max(max_len, len(str(cell.value or '')))

                # Apply styles based on row placement
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                else:
                    cell.fill = data_fill

            # Apply widths dynamically
            worksheet.column_dimensions[col_letter].width = max(max_len + 5, 25)

    print(f"Spreadsheet generated securely: {file_name}")
    print("--- Data Pipeline Finished ---")
