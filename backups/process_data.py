import os
import httpx
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

print("--- Starting Premium DaaS Data Pipeline (Header Inspection Mode) ---")

def fetch_raw_api_sample(city_name: str):
    """
    Fetches raw property objects directly from the live RentCast API.
    """
    target_city = city_name.strip() if city_name else "Dallas"
    print(f"Connecting to live database stream to inspect data headers for: {target_city}...")

    # Targeted active real estate data stream path
    # end points info page on rentcast: https://developers.rentcast.io/reference/rental-listings-long-term
    url = "https://api.rentcast.io/v1/listings/rental/long-term"
      # "https://api.rentcast.io/v1/listings/active"

    query_params = {
        "city": target_city,
        "state": "TX",
        "status": "Active",
        "limit": 5  # We only need a few samples to inspect the column names
    }

    headers = {
        "Accept": "application/json",
        "X-Api-Key": os.getenv("RENTCAST_REAL_ESTATE_API_KEY")
    }

    try:
        response = httpx.get(url, headers=headers, params=query_params, timeout=20)
        if response.status_code == 200:
            return response.json()
        else:
            print(f"API Warning: Status {response.status_code}. Details: {response.text}")
            return []
    except Exception as e:
        print(f"Network error: {e}")
        return []


def extract_raw_headers(raw_properties):
    """
    Inspected data catcher: Unwraps the RentCast API dictionary safely.
    """
    # Safely unbox JSON collections or nested structural data keys
    if isinstance(raw_properties, dict):
        print(f"API returned a Dictionary structure. Available top-level keys: {list(raw_properties.keys())}")

        # Extract listings array layer out of standard API wrapper structures
        actual_list = raw_properties.get("listings") or raw_properties.get("data") or []

        if isinstance(actual_list, list) and len(actual_list) > 0:
            raw_properties = actual_list
        else:
            # Standard structural failover fallback mapping
            if len(raw_properties) > 0:
                raw_properties = [raw_properties]
            else:
                raw_properties = []

    if not raw_properties or not isinstance(raw_properties, list) or len(raw_properties) == 0:
        print("Data stream empty or malformed. Ensure your key is active in GitHub Secrets.")
        return pd.DataFrame([{"Detected API Response Keys": "ERROR: Vault credentials or data pull returned empty."}])

    # Grab the very first raw property dictionary object from the list
    first_house_sample = raw_properties[0]

    # Extract all data keys
    api_header_names = list(first_house_sample.keys())
    print(f"SUCCESS: Successfully isolated {len(api_header_names)} unique database headers!")

    header_tracking_rows = []
    for header in api_header_names:
        sample_value = str(first_house_sample.get(header, "Empty/Null"))
        header_tracking_rows.append({
            "Available API Column Keys": header,
            "Sample Raw Data Example": sample_value[:50]
        })

    return pd.DataFrame(header_tracking_rows)


if __name__ == "__main__":
    input_location = os.getenv("FILTER_TEXT", "Dallas")

    # 1. Harvest raw data records
    raw_sample = fetch_raw_api_sample(input_location)

    # 2. Extract the exact database header labels into a DataFrame
    headers_df = extract_raw_headers(raw_sample)

    file_name = "Premium_Undervalued_Deals.xlsx"

    # 3. Save the header inspection report to your repository spreadsheet
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        headers_df.to_excel(writer, sheet_name="API Keys Blueprint", index=False)

        workbook = writer.book
        worksheet = writer.sheets["API Keys Blueprint"]

        # Style the inspection report
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")  # Purple Blueprint mode
        data_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")

        for col_num, column_title in enumerate(headers_df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        for row in worksheet.iter_rows(min_row=2, max_row=len(headers_df) + 1, min_col=1, max_col=2):
            for cell in row:
                cell.fill = data_fill

        # Auto-adjust column dimensions cleanly using explicit openpyxl cell properties
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 20)

    print(f"Spreadsheet generated securely: {file_name}")
    print("--- Data Pipeline Finished ---")
